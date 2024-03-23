import io
import os
import time
from flask import Flask, request, jsonify , send_file
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
from openpyxl import load_workbook
from utils import Utils
from datetime import datetime
from googleapiclient.http import MediaFileUpload ,MediaIoBaseDownload
import shutil
import pandas as pd
from flask_cors import CORS
from google.cloud.firestore_v1 import Increment, Transaction
from dotenv import load_dotenv


load_dotenv()

firebase_credentials = {
    "type": os.getenv("type"),
    "project_id": os.getenv("project_id"),
    "private_key_id": os.getenv("private_key_id"),
    "private_key": os.getenv("private_key").replace('\\n', '\n'),  # Replace escaped newlines
    "client_email": os.getenv("client_email"),
    "client_id": os.getenv("client_id"),
    "auth_uri": os.getenv("auth_uri"),
    "token_uri": os.getenv("token_uri"),
    "auth_provider_x509_cert_url": os.getenv("auth_provider_x509_cert_url"),
    "client_x509_cert_url": os.getenv("client_x509_cert_url"),
    "universe_domain": os.getenv("universe_domain"),
}



class GoogleDriveAPI:
    def __init__(self, creds):
        self.app = Flask(__name__)
        self.creds = creds
        self.service = self.build_service()
        self.db = self.initialize_firebase()
        self.original_invoice_file_path = ''  

        @self.app.route('/', methods=['GET'])
        def check_health():
            return 'health OK'
        
        @self.app.route('/get_folder', methods=['GET'])
        def getfolder():
            folder_name = request.args.get('foldername')
            if not folder_name:
                return jsonify({"error": "Folder name is required"})
            
            try:
              folder_items = Utils.get_folder(folder_name,self.service)
              
              if not folder_items:
                  print(f"Folder '{folder_name}' not found.")
                  return jsonify({"error": f"Folder '{folder_name}' not found."})
              
              folder_id = folder_items[0]["id"]

              results = (
                  self.service.files() 
                  .list(q=f"'{folder_id}' in parents", pageSize=100, fields="nextPageToken, files(id, name)")
                  .execute()
              )
              items = results.get("files", [])

              return jsonify(items)
            except HttpError as error:
              print(f"An error occurred: {error}")
              return jsonify({"error": f"An error occurred: {error}"})
            
        @self.app.route('/download', methods=['POST'])
        def download():
            file_id = request.args.get('id')
            try:
                # Retrieve file metadata to get the filename
                file_metadata = self.service.files().get(fileId=file_id, fields='name').execute()
                filename = file_metadata['name']
                data = self.service.files().get_media(fileId=file_id)
                file = io.BytesIO()
                downloader = MediaIoBaseDownload(file, data)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                    print(f"Download {int(status.progress() * 100)}.")

                # Seek to the beginning of the BytesIO object
                file.seek(0)
                return send_file(
                            file,
                            as_attachment=True,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            download_name=filename+'xlxs' # Change the filename as needed
                        )
            except HttpError as error:
                print(f"An error occurred: {error}")
                return "An error occurred while downloading the file", 500      


        @self.app.route('/delete', methods=['POST'])
        def delete():
            file_id = request.args.get('id')
            try:
                # Call Google Drive API to delete the file
                self.service.files().delete(fileId=file_id).execute()
        # Return a success message
                return jsonify({'message': 'File deleted successfully'}), 200
            except HttpError as error:
                # Handle errors
                print(f"An error occurred: {error}")
                return jsonify({'error': 'Failed to delete file'}), 500


        @self.app.route('/upload_excel', methods=['POST'])
        async def upload():

            data = request.get_json()

            if not data:
                return jsonify({"error": "No JSON data provided"})
            
            self.original_invoice_file_path = data.get('type')+'.xlsx'

            collection_name = "Autoinvoice"

            docs = self.db.collection(collection_name).get()

            latest = 0

            for doc in docs:
                latest = doc.get(data.get('type'))
                
            new_file = data.get('for').get('company').get('name')+'.xlsx'

            # Create a copy of the original file
            shutil.copyfile(self.original_invoice_file_path, new_file)

            client =  data.get('for').get('company').get('name') + " (UEN:" + str(data.get('for').get('company').get('uen')) + ")"

            A = load_workbook(new_file)
            B = A['Quote']
            B.cell(row =17,column=6,value=latest)
            B.cell(row =12,column=2,value=data.get('for').get('name'))
            B.cell(row =13,column=2,value=client)
            B.cell(row =14,column=2,value=data.get('for').get('email'))
            B.cell(row =15,column=2,value=data.get('for').get('phone'))
            B.cell(row =14,column=7,value=data.get('to').get('email'))
            B.cell(row =15,column=7,value=data.get('to').get('phone'))
            B.cell(row =17,column=7,value=data.get('dueon'))

            items = data.get('items')
            for index, item in enumerate(items, start=21):
                B.cell(row=index , column=2, value=item.get('desc'))
                B.cell(row=index , column=7, value=int(item.get('price')))

            A.save(new_file)
            A.close()

            file_metadata = {
                'name': client ,
                'parents': [data.get('tofolder')],  # Specify the parent folder ID
            }

            try:
                uploaded_file = await self.upload_file(new_file,file_metadata)
                if uploaded_file:
                    try:
                        os.remove(new_file)
                        # Increment the latest value by 1
                        latest += 1
                        # Update the Firestore document
                        doc_ref = self.db.collection(collection_name).document(collection_name)
                        doc_ref.update({data.get('type'): latest})
                        # Increment the value of after removing the file doc.get(data.get('type')) in docs in the firebase here after deleting
                        return jsonify({"message": f'File "{new_file}" successfully uploaded'})
                    except PermissionError as e:
                        return jsonify({"message": "permission error"})
                    except FileNotFoundError as e:
                        return jsonify({"message": "file error"})
                  
                else:
                    print('Failed to retrieve file ID. File may not have been uploaded successfully.')
                    return jsonify({"error": "error test"})
                    
            except HttpError as error:
                print(f"An error occurred: {error}")



    async def upload_file(self,new_file,file_metadata):
        media = MediaFileUpload(new_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        uploaded_file = self.service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        return uploaded_file

    def build_service(self):
        return build("drive", "v3", credentials=self.creds)
    
    def initialize_firebase(self):
        if not firebase_admin._apps:
            cred = credentials.Certificate(firebase_credentials)
            firebase_admin.initialize_app(cred , name='Autoinvoice')
        return firestore.client(app=firebase_admin.get_app(name='Autoinvoice'))


    def run(self):
        CORS(self.app)
        self.app.run(port=5000)

if __name__ == '__main__':
    app = GoogleDriveAPI(creds=Utils.load_credentials())
    app.run()
